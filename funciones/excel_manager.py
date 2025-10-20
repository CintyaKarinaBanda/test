from openpyxl import load_workbook

#Funcion para escribir tablas grandes
def write_report(data, start_row=1, start_col=1, target_filename="CheckList.xlsx"):
    wb = load_workbook(target_filename)
    ws = wb.active

    print(f"DEBUG write_report: Escribiendo {len(data)} filas desde fila {start_row}, columna {start_col}")
    
    for i, row_data in enumerate(data, start=start_row):
        print(f"DEBUG: Escribiendo fila {i} con {len(row_data)} valores")
        for j, value in enumerate(row_data, start=start_col):
            try:
                print(f"DEBUG: Celda ({i},{j}) = {value}")
                ws.cell(row=i, column=j, value=value)
            except Exception as e:
                print(f"ERROR: No se pudo escribir en celda ({i},{j}): {e}")
                continue

    wb.save(target_filename)
    print(f"Reporte guardado como: {target_filename}")

#Funcion para escribir en una sola columna
def write_column(comments, start_row=1, start_col=1, target_filename="CheckList.xlsx"):
    wb = load_workbook(target_filename)
    ws = wb.active

    for i, comment in enumerate(comments, start=start_row):
        ws.cell(row=i, column=start_col, value=comment)

    wb.save(target_filename)
    print(f"Comentarios guardados en: {target_filename}")

#Funcion especializada para dashboard de RDS en el Excel
def write_rds_dashboard_metrics(data, start_row=1, start_col=1, target_filename="CheckList.xlsx"):
    wb = load_workbook(target_filename)
    ws = wb.active

    for i, row_data in enumerate(data, start=start_row):
        for j, value in enumerate(row_data, start=start_col):
            ws.cell(row=i, column=j, value=value)

    wb.save(target_filename)
    print(f"Datos del dashboard de RDS guardados en: {target_filename}")